#############################
# main.py                   #
#                           #
# GUI that supports         #
# functionality of showing  #
# finance data through      #
# graphs, charts, and       #
# spreadsheets. Queries and #
# fetches data from MS SQL  #
# server.                   #
#############################

###IMPORTS###
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QFont
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas 
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
import openpyxl
import query_data
from add_new import AnotherWindow


# main window
class Window(QWidget):
    ''' Main GUI Window for Feeble Finance. Includes GUI methods as well as 
    matplotlib.pyplot methods to show data derived from query_data functions.

    Methods:

    self.tab1 - The PyQt5 UI elements for title and description page.
    self.tab2 - The PyQt5 UI elements and three matplotlib.pyplot canvases for 
                the graphs and charts page.
    self.tab3 - The PyQt5 UI elements for the personal data page.
    self.layoutUI - The UI that connects the three tabs together and creates the 
                    main interface.
    self.refresh - Calls self.plot, self.chart, and query_data.get_top_three() to 
                    refresh UI based on current financial data.
    self.plot - Creates a financial line graph for tab 2. 
    self.chart - Creates the pie and bar graphs for tab 2. 
    self.add_form - Functionality for the 'Add More Data' button on tab 3. Calls 
                        AnotherWindow() class from add_new.py.
    self.excel_sheet - Creates the excel sheet displaying financial data for tab 3. 
    self.closeEvent - Calls query_data.finish() upon closing the GUI so that the 
                        MS SQL connection is stopped. 
    '''
      
    # constructor
    def __init__(self, parent=None):
        super(Window, self).__init__(parent=parent)
        self.figure = plt.figure() # plot
        self.canvas = FigureCanvas(self.figure)
        self.figure2 = plt.figure()
        self.canvas2 = FigureCanvas(self.figure2)
        self.figure3 = plt.figure()
        self.canvas3 = FigureCanvas(self.figure3)
        self.high_expenses = QLabel()
        self.table_widget = QTableWidget()
        self.layoutUI()
    
    # "Title Screen" tab
    def tab1(self): 
        # GUI items
        tab1_items = QWidget()
        tab1_items.layout = QVBoxLayout(tab1_items) 
        l = QLabel() 
        l.setText("Feeble Finance") 
        l.setFont(QFont('Arial', 30))
        l1 = QLabel()
        l1.setText("Track and visualize your monthly finances.\n\nTry adding some monthly finances in 'My Data', and click 'Refresh Data' button in 'Graphs and Charts'!\n\n\n\n\n\n\n@Jasmine Wallin 2024")
        
        # adding items onto the layout
        tab1_items.layout.addWidget(l) 
        tab1_items.layout.addWidget(l1)
        tab1_items.setLayout(tab1_items.layout) 

        return tab1_items

    # Tab to show data through graphs and charts
    def tab2(self):
        # GUI items
        tab2_items = QWidget()
        tab2_items.layout = QVBoxLayout(tab2_items)
        Button_01 = QPushButton("Refresh Data")
        Button_01.clicked.connect(self.refresh)
        tab2_items.layout.addWidget(Button_01)
        tab2_items.setLayout(tab2_items.layout)

        # creating a scrollable widget
        scroll = QScrollArea(self)
        tab2_items.layout.addWidget(scroll)
        scroll.setWidgetResizable(True)
        scrollContent = QWidget(scroll)
        scrollLayout = QVBoxLayout(scrollContent)
        scrollContent.setLayout(scrollLayout)

        # Defining items that will go on screen
        graph_text = QLabel() 
        graph_text.setText("Graph of Expenses") 
        chart_text = QLabel() 
        chart_text.setText("Charts of Expenses")   

        # frames for canvasses
        frame = QFrame()
        lay = QVBoxLayout(frame)
        lay.addWidget(self.canvas)
        frame2 = QFrame()
        lay2 = QHBoxLayout(frame2)
        lay2.addWidget(self.canvas2)
        lay2.addWidget(self.canvas3)

        # Adding items to the screen through a vertical layout
        scrollLayout.addWidget(graph_text)
        scrollLayout.addWidget(frame)
        scrollLayout.addWidget(chart_text)
        scrollLayout.addWidget(frame2) 
        scrollLayout.addWidget(self.high_expenses)
        
        scroll.setWidget(scrollContent)
        
        return tab2_items

    # Tab to show current database, and add more items to the database
    def tab3(self):
        # GUI items
        tab3_items = QWidget()
        tab3_items.layout = QVBoxLayout(tab3_items)
        current_text = QLabel() 
        current_text.setText("My Current Financial Data") 
        Button_ref = QPushButton("Refresh Data")
        Button_ref.clicked.connect(self.excel_sheet)
        add_button = QPushButton("Add More Data") 
        add_button.clicked.connect(self.add_form)
 
        # adding items onto the layout
        tab3_items.layout.addWidget(Button_ref)
        tab3_items.layout.addWidget(current_text)
        tab3_items.layout.addWidget(self.table_widget)
        tab3_items.layout.addWidget(add_button)
        tab3_items.setLayout(tab3_items.layout)

        return tab3_items

    # main layout with bar of tabs
    def layoutUI(self):

        self.principle_layout = QVBoxLayout(self) 
  
        # Initialize tab screen 
        self.tabs = QTabWidget() 
        self.tab1 = self.tab1()
        self.tab2 = self.tab2()
        self.tab3 = self.tab3() 
        self.tabs.resize(300, 200) 
  
        # Add tabs 
        self.tabs.addTab(self.tab1, "Home") 
        self.tabs.addTab(self.tab2, "Graphs and Charts") 
        self.tabs.addTab(self.tab3, "My Data") 

        # Add tabs to widget 
        self.principle_layout.addWidget(self.tabs) 
        self.setLayout(self.principle_layout)

    # refresh plot and chart
    def refresh(self):
        # refreshing the plot for expense over time
        self.plot()
        # refreshing the monthly chart expense visuals
        self.chart()
        # refreshing the highest expenses text
        top1, top2, top3 = query_data.find_top_three()
        highstring = "Your Current Highest Expenses:\n*" + top1 + "\n*" + top2 + "\n*" + top3
        self.high_expenses.setText(highstring)

    # Financial graph for tab 3 
    def plot(self):
        self.figure.clear()
        ax = self.figure.add_subplot(111) # create an axis

        # adding labels
        ax.set_ylabel('Amount Per Month')
        ax.set_xlabel('Date')
        ax.set_title('Expense Breakdown Over Time')

        ### plot data lines ###
        # general data
        general_data = query_data.graph_data('')
        # needs data
        needs_data = query_data.graph_data('needs')
        # fufillment data
        fuf_data = query_data.graph_data('fufillment')
        # social data
        soc_data = query_data.graph_data('social')
        # extra data
        extra_data = query_data.graph_data('extra')
        # plot everything
        if general_data != ([], []): # if there is data to display yet
            ax.plot(general_data[0], general_data[1], needs_data[0], needs_data[1], fuf_data[0], fuf_data[1],soc_data[0], soc_data[1],extra_data[0], extra_data[1],'b-') 
            ax.legend(('total', 'needs', 'fufillment', 'social', 'extra'))
  
        # refresh canvas
        self.canvas.draw()

    # Charts for tab 3
    def chart(self):

        # getting the chart data for both graphs
        new = False
        chart_data = query_data.chart_data()
        if chart_data == [0.0, 0.0, 0.0, 0.0]: # if there is no data yet
            new = True

        ### chart 1 ###
        self.figure2.clear()
        ax = self.figure2.add_subplot(111)
        ax.set_title('Current Monthly Expense Breakdown')
        if not new:
            ax.pie(chart_data, labels=['needs', 'fufillment', 'social', 'extra'])

        ### chart 2 ###
        self.figure3.clear()
        ax2 = self.figure3.add_subplot(111)
        ax2.set_title('Current Monthly Expense Breakdown')
        ax2.set_xlabel("Type of Expense")
        ax2.set_ylabel("Amount Spent Monthly")
        if not new:
            ax2.bar(['needs', 'fufillment', 'social', 'extra'], chart_data)
        # plotting / drawing charts
        self.canvas2.draw()
        self.canvas3.draw()
    
    # finds input from a form to add to data
    def add_form(self):
        self.w = AnotherWindow()
        self.w.show()
    
    # creates the GUI data sheet and the data inside
    def excel_sheet(self):
        # initialising the finance_data file
        query_data.make_excel() # writing the finance_data.xlsx file
        path = "finance_data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        # row and col count
        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)
        
        # column headers
        list_values = list(sheet.values)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])
        
        # filling the table
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index , col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

    # close the query connection upon exiting the program
    def closeEvent(self, event):
        query_data.finish()
  
# driver code
if __name__ == '__main__':
    
    # creating the SQL database if it doesn't exist already
    query_data.create_database()

    # creating apyqt5 application
    app = QApplication(sys.argv)
  
    # creating a window object
    main = Window()
    main.setWindowTitle("Feeble Finance")
    main.setGeometry(200, 200, 1200, 760)

    # showing the window
    main.show()

    # loop
    sys.exit(app.exec_())
